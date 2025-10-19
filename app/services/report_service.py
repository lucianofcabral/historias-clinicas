"""
Servicio de generación de reportes y exportación
Soporta PDF y Excel
"""

from pathlib import Path
from datetime import datetime, date
from typing import List, Optional
import io

# PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Database
from sqlmodel import Session, select
from app.database import engine
from app.models.patient import Patient
from app.models.consultation import Consultation
from app.models.medical_study import MedicalStudy


class ReportService:
    """Servicio para generar reportes en PDF y Excel"""

    @staticmethod
    def generate_patient_history_pdf(patient_id: int) -> bytes:
        """
        Genera un PDF con el historial completo de un paciente

        Args:
            patient_id: ID del paciente

        Returns:
            bytes: Contenido del PDF
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Estilos personalizados
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#2563eb"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#1e40af"),
            spaceAfter=12,
            spaceBefore=12,
        )

        # Obtener datos
        with Session(engine) as session:
            patient = session.get(Patient, patient_id)
            if not patient:
                raise ValueError("Paciente no encontrado")

            consultations = session.exec(
                select(Consultation)
                .where(Consultation.patient_id == patient_id)
                .order_by(Consultation.consultation_date.desc())
            ).all()

            studies = session.exec(
                select(MedicalStudy)
                .where(MedicalStudy.patient_id == patient_id)
                .order_by(MedicalStudy.study_date.desc())
            ).all()

        # Título
        story.append(Paragraph("Historia Clínica Completa", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Información del paciente
        story.append(Paragraph("Datos del Paciente", heading_style))

        patient_data = [
            ["Nombre Completo:", f"{patient.first_name} {patient.last_name}"],
            ["DNI:", patient.dni or "N/A"],
            [
                "Fecha de Nacimiento:",
                patient.birth_date.strftime("%d/%m/%Y") if patient.birth_date else "N/A",
            ],
            ["Género:", patient.gender or "N/A"],
            ["Tipo de Sangre:", patient.blood_type or "N/A"],
            ["Teléfono:", patient.phone or "N/A"],
            ["Email:", patient.email or "N/A"],
            ["Dirección:", patient.address or "N/A"],
        ]

        patient_table = Table(patient_data, colWidths=[2 * inch, 4 * inch])
        patient_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e0e7ff")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                    ("ALIGN", (1, 0), (1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        story.append(patient_table)
        story.append(Spacer(1, 0.3 * inch))

        # Antecedentes médicos
        if patient.allergies or patient.chronic_conditions or patient.family_history:
            story.append(Paragraph("Antecedentes Médicos", heading_style))

            antecedentes_data = []
            if patient.allergies:
                antecedentes_data.append(["Alergias:", patient.allergies])
            if patient.chronic_conditions:
                antecedentes_data.append(["Condiciones Crónicas:", patient.chronic_conditions])
            if patient.family_history:
                antecedentes_data.append(["Antecedentes Familiares:", patient.family_history])

            if antecedentes_data:
                antecedentes_table = Table(antecedentes_data, colWidths=[2 * inch, 4 * inch])
                antecedentes_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#fef3c7")),
                            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                            ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                            ("ALIGN", (1, 0), (1, -1), "LEFT"),
                            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 8),
                            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ]
                    )
                )
                story.append(antecedentes_table)
                story.append(Spacer(1, 0.3 * inch))

        # Consultas
        if consultations:
            story.append(PageBreak())
            story.append(Paragraph(f"Consultas ({len(consultations)} registros)", heading_style))
            story.append(Spacer(1, 0.2 * inch))

            for i, consultation in enumerate(consultations, 1):
                # Encabezado de consulta
                consult_header = (
                    f"Consulta #{i} - {consultation.consultation_date.strftime('%d/%m/%Y')}"
                )
                story.append(Paragraph(consult_header, styles["Heading3"]))

                consult_data = [
                    ["Motivo:", consultation.reason or "N/A"],
                    ["Síntomas:", consultation.symptoms or "N/A"],
                    ["Diagnóstico:", consultation.diagnosis or "N/A"],
                    ["Tratamiento:", consultation.treatment or "N/A"],
                ]

                # Signos vitales
                vitals = []
                if consultation.blood_pressure:
                    vitals.append(f"PA: {consultation.blood_pressure}")
                if consultation.heart_rate:
                    vitals.append(f"FC: {consultation.heart_rate} lpm")
                if consultation.temperature:
                    vitals.append(f"T°: {consultation.temperature}°C")
                if consultation.weight:
                    vitals.append(f"Peso: {consultation.weight} kg")
                if consultation.height:
                    vitals.append(f"Altura: {consultation.height} cm")

                if vitals:
                    consult_data.append(["Signos Vitales:", " | ".join(vitals)])

                if consultation.notes:
                    consult_data.append(["Notas:", consultation.notes])

                if consultation.next_visit:
                    consult_data.append(
                        ["Próxima Visita:", consultation.next_visit.strftime("%d/%m/%Y")]
                    )

                consult_table = Table(consult_data, colWidths=[1.5 * inch, 4.5 * inch])
                consult_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#dbeafe")),
                            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                            ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                            ("ALIGN", (1, 0), (1, -1), "LEFT"),
                            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                            ("TOPPADDING", (0, 0), (-1, -1), 6),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ]
                    )
                )
                story.append(consult_table)
                story.append(Spacer(1, 0.2 * inch))

        # Estudios médicos
        if studies:
            story.append(PageBreak())
            story.append(Paragraph(f"Estudios Médicos ({len(studies)} registros)", heading_style))
            story.append(Spacer(1, 0.2 * inch))

            for i, study in enumerate(studies, 1):
                study_header = f"Estudio #{i} - {study.study_type}"
                story.append(Paragraph(study_header, styles["Heading3"]))

                study_data = [
                    ["Nombre:", study.study_name or "N/A"],
                    [
                        "Fecha:",
                        study.study_date.strftime("%d/%m/%Y") if study.study_date else "N/A",
                    ],
                    ["Institución:", study.institution or "N/A"],
                    ["Médico Solicitante:", study.requesting_doctor or "N/A"],
                    ["Resultados:", study.results or "N/A"],
                ]

                if study.observations:
                    study_data.append(["Observaciones:", study.observations])

                if study.diagnosis:
                    study_data.append(["Diagnóstico:", study.diagnosis])

                # Estado
                status_items = []
                if study.is_pending:
                    status_items.append("Pendiente")
                if study.is_critical:
                    status_items.append("Crítico")
                if study.requires_followup:
                    status_items.append("Requiere Seguimiento")

                if status_items:
                    study_data.append(["Estado:", " | ".join(status_items)])

                study_table = Table(study_data, colWidths=[1.5 * inch, 4.5 * inch])
                study_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#fce7f3")),
                            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                            ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                            ("ALIGN", (1, 0), (1, -1), "LEFT"),
                            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                            ("TOPPADDING", (0, 0), (-1, -1), 6),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ]
                    )
                )
                story.append(study_table)
                story.append(Spacer(1, 0.2 * inch))

        # Pie de página con fecha de generación
        story.append(Spacer(1, 0.5 * inch))
        footer_text = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        story.append(
            Paragraph(
                footer_text,
                ParagraphStyle(
                    "Footer",
                    parent=styles["Normal"],
                    fontSize=8,
                    textColor=colors.grey,
                    alignment=TA_RIGHT,
                ),
            )
        )

        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_patient_history_excel(patient_id: int) -> bytes:
        """
        Genera un archivo Excel con el historial completo de un paciente

        Args:
            patient_id: ID del paciente

        Returns:
            bytes: Contenido del archivo Excel
        """
        wb = Workbook()

        # Estilos
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Obtener datos
        with Session(engine) as session:
            patient = session.get(Patient, patient_id)
            if not patient:
                raise ValueError("Paciente no encontrado")

            consultations = session.exec(
                select(Consultation)
                .where(Consultation.patient_id == patient_id)
                .order_by(Consultation.consultation_date.desc())
            ).all()

            studies = session.exec(
                select(MedicalStudy)
                .where(MedicalStudy.patient_id == patient_id)
                .order_by(MedicalStudy.study_date.desc())
            ).all()

        # Hoja 1: Información del Paciente
        ws_patient = wb.active
        ws_patient.title = "Datos del Paciente"

        # Título
        ws_patient["A1"] = "HISTORIA CLÍNICA"
        ws_patient["A1"].font = Font(bold=True, size=16, color="2563eb")
        ws_patient.merge_cells("A1:B1")

        ws_patient["A2"] = f"Paciente: {patient.first_name} {patient.last_name}"
        ws_patient["A2"].font = Font(bold=True, size=12)
        ws_patient.merge_cells("A2:B2")

        # Datos del paciente
        row = 4
        patient_data = [
            ("DNI", patient.dni or "N/A"),
            (
                "Fecha de Nacimiento",
                patient.birth_date.strftime("%d/%m/%Y") if patient.birth_date else "N/A",
            ),
            ("Género", patient.gender or "N/A"),
            ("Tipo de Sangre", patient.blood_type or "N/A"),
            ("Teléfono", patient.phone or "N/A"),
            ("Email", patient.email or "N/A"),
            ("Dirección", patient.address or "N/A"),
            ("Alergias", patient.allergies or "N/A"),
            ("Condiciones Crónicas", patient.chronic_conditions or "N/A"),
            ("Antecedentes Familiares", patient.family_history or "N/A"),
            ("Notas", patient.notes or "N/A"),
        ]

        for label, value in patient_data:
            ws_patient[f"A{row}"] = label
            ws_patient[f"B{row}"] = value
            ws_patient[f"A{row}"].font = Font(bold=True)
            ws_patient[f"A{row}"].fill = subheader_fill
            ws_patient[f"A{row}"].border = border
            ws_patient[f"B{row}"].border = border
            ws_patient[f"B{row}"].alignment = Alignment(wrap_text=True)
            row += 1

        # Ajustar anchos
        ws_patient.column_dimensions["A"].width = 25
        ws_patient.column_dimensions["B"].width = 50

        # Hoja 2: Consultas
        if consultations:
            ws_consult = wb.create_sheet("Consultas")

            # Encabezados
            headers = [
                "Fecha",
                "Motivo",
                "Síntomas",
                "Diagnóstico",
                "Tratamiento",
                "Presión",
                "FC",
                "Temp",
                "Peso",
                "Altura",
                "Notas",
                "Próxima Visita",
            ]

            for col, header in enumerate(headers, 1):
                cell = ws_consult.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Datos
            for row, consultation in enumerate(consultations, 2):
                ws_consult.cell(
                    row,
                    1,
                    consultation.consultation_date.strftime("%d/%m/%Y")
                    if consultation.consultation_date
                    else "",
                )
                ws_consult.cell(row, 2, consultation.reason or "")
                ws_consult.cell(row, 3, consultation.symptoms or "")
                ws_consult.cell(row, 4, consultation.diagnosis or "")
                ws_consult.cell(row, 5, consultation.treatment or "")
                ws_consult.cell(row, 6, consultation.blood_pressure or "")
                ws_consult.cell(row, 7, consultation.heart_rate or "")
                ws_consult.cell(row, 8, consultation.temperature or "")
                ws_consult.cell(row, 9, consultation.weight or "")
                ws_consult.cell(row, 10, consultation.height or "")
                ws_consult.cell(row, 11, consultation.notes or "")
                ws_consult.cell(
                    row,
                    12,
                    consultation.next_visit.strftime("%d/%m/%Y") if consultation.next_visit else "",
                )

                # Aplicar bordes y alineación
                for col in range(1, 13):
                    cell = ws_consult.cell(row, col)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Ajustar anchos
            ws_consult.column_dimensions["A"].width = 12
            ws_consult.column_dimensions["B"].width = 20
            ws_consult.column_dimensions["C"].width = 25
            ws_consult.column_dimensions["D"].width = 25
            ws_consult.column_dimensions["E"].width = 25
            ws_consult.column_dimensions["F"].width = 10
            ws_consult.column_dimensions["G"].width = 8
            ws_consult.column_dimensions["H"].width = 8
            ws_consult.column_dimensions["I"].width = 8
            ws_consult.column_dimensions["J"].width = 8
            ws_consult.column_dimensions["K"].width = 30
            ws_consult.column_dimensions["L"].width = 12

        # Hoja 3: Estudios
        if studies:
            ws_studies = wb.create_sheet("Estudios Médicos")

            # Encabezados
            headers = [
                "Fecha",
                "Tipo",
                "Nombre",
                "Institución",
                "Médico",
                "Resultados",
                "Observaciones",
                "Diagnóstico",
                "Estado",
            ]

            for col, header in enumerate(headers, 1):
                cell = ws_studies.cell(1, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Datos
            for row, study in enumerate(studies, 2):
                status = []
                if study.is_pending:
                    status.append("Pendiente")
                if study.is_critical:
                    status.append("Crítico")
                if study.requires_followup:
                    status.append("Seguimiento")

                ws_studies.cell(
                    row, 1, study.study_date.strftime("%d/%m/%Y") if study.study_date else ""
                )
                ws_studies.cell(row, 2, study.study_type or "")
                ws_studies.cell(row, 3, study.study_name or "")
                ws_studies.cell(row, 4, study.institution or "")
                ws_studies.cell(row, 5, study.requesting_doctor or "")
                ws_studies.cell(row, 6, study.results or "")
                ws_studies.cell(row, 7, study.observations or "")
                ws_studies.cell(row, 8, study.diagnosis or "")
                ws_studies.cell(row, 9, " | ".join(status) if status else "")

                # Aplicar bordes y alineación
                for col in range(1, 10):
                    cell = ws_studies.cell(row, col)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Resaltar estudios críticos
                if study.is_critical:
                    for col in range(1, 10):
                        ws_studies.cell(row, col).fill = PatternFill(
                            start_color="fee2e2", end_color="fee2e2", fill_type="solid"
                        )

            # Ajustar anchos
            ws_studies.column_dimensions["A"].width = 12
            ws_studies.column_dimensions["B"].width = 15
            ws_studies.column_dimensions["C"].width = 20
            ws_studies.column_dimensions["D"].width = 20
            ws_studies.column_dimensions["E"].width = 20
            ws_studies.column_dimensions["F"].width = 30
            ws_studies.column_dimensions["G"].width = 30
            ws_studies.column_dimensions["H"].width = 30
            ws_studies.column_dimensions["I"].width = 15

        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_consultations_report_pdf(
        start_date: Optional[date] = None, end_date: Optional[date] = None
    ) -> bytes:
        """
        Genera un reporte PDF de consultas en un rango de fechas

        Args:
            start_date: Fecha inicio (opcional)
            end_date: Fecha fin (opcional)

        Returns:
            bytes: Contenido del PDF
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=20,
            textColor=colors.HexColor("#2563eb"),
            spaceAfter=20,
            alignment=TA_CENTER,
        )

        # Obtener consultas
        with Session(engine) as session:
            query = select(Consultation).order_by(Consultation.consultation_date.desc())

            if start_date:
                query = query.where(Consultation.consultation_date >= start_date)
            if end_date:
                query = query.where(Consultation.consultation_date <= end_date)

            consultations = session.exec(query).all()

            # Obtener pacientes
            patient_ids = {c.patient_id for c in consultations}
            patients = {
                p.id: p
                for p in session.exec(select(Patient).where(Patient.id.in_(patient_ids))).all()
            }

        # Título
        title = "Reporte de Consultas"
        if start_date and end_date:
            title += f"\n{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        elif start_date:
            title += f"\nDesde {start_date.strftime('%d/%m/%Y')}"
        elif end_date:
            title += f"\nHasta {end_date.strftime('%d/%m/%Y')}"

        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.3 * inch))

        # Estadísticas
        story.append(Paragraph(f"Total de Consultas: {len(consultations)}", styles["Heading2"]))
        story.append(Spacer(1, 0.2 * inch))

        # Tabla de consultas
        if consultations:
            data = [["Fecha", "Paciente", "Motivo", "Diagnóstico"]]

            for consultation in consultations:
                patient = patients.get(consultation.patient_id)
                patient_name = f"{patient.first_name} {patient.last_name}" if patient else "N/A"

                data.append(
                    [
                        consultation.consultation_date.strftime("%d/%m/%Y"),
                        patient_name,
                        consultation.reason[:50] + "..."
                        if consultation.reason and len(consultation.reason) > 50
                        else consultation.reason or "",
                        consultation.diagnosis[:50] + "..."
                        if consultation.diagnosis and len(consultation.diagnosis) > 50
                        else consultation.diagnosis or "",
                    ]
                )

            table = Table(data, colWidths=[1.2 * inch, 2 * inch, 2.3 * inch, 2.3 * inch])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 11),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_studies_report_excel(
        study_type: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """
        Genera un reporte Excel de estudios médicos

        Args:
            study_type: Tipo de estudio (opcional)
            start_date: Fecha inicio (opcional)
            end_date: Fecha fin (opcional)

        Returns:
            bytes: Contenido del archivo Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Estudios Médicos"

        # Estilos
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        critical_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Obtener estudios
        with Session(engine) as session:
            query = select(MedicalStudy).order_by(MedicalStudy.study_date.desc())

            if study_type:
                query = query.where(MedicalStudy.study_type == study_type)
            if start_date:
                query = query.where(MedicalStudy.study_date >= start_date)
            if end_date:
                query = query.where(MedicalStudy.study_date <= end_date)

            studies = session.exec(query).all()

            # Obtener pacientes
            patient_ids = {s.patient_id for s in studies}
            patients = {
                p.id: p
                for p in session.exec(select(Patient).where(Patient.id.in_(patient_ids))).all()
            }

        # Título
        ws["A1"] = "REPORTE DE ESTUDIOS MÉDICOS"
        ws["A1"].font = Font(bold=True, size=16, color="2563eb")
        ws.merge_cells("A1:I1")

        if study_type or start_date or end_date:
            filters = []
            if study_type:
                filters.append(f"Tipo: {study_type}")
            if start_date:
                filters.append(f"Desde: {start_date.strftime('%d/%m/%Y')}")
            if end_date:
                filters.append(f"Hasta: {end_date.strftime('%d/%m/%Y')}")

            ws["A2"] = " | ".join(filters)
            ws["A2"].font = Font(size=10, italic=True)
            ws.merge_cells("A2:I2")

        # Encabezados
        headers = [
            "Fecha",
            "Paciente",
            "Tipo",
            "Nombre",
            "Institución",
            "Médico",
            "Resultados",
            "Diagnóstico",
            "Estado",
        ]

        header_row = 4 if (study_type or start_date or end_date) else 3

        for col, header in enumerate(headers, 1):
            cell = ws.cell(header_row, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Datos
        for row_idx, study in enumerate(studies, header_row + 1):
            patient = patients.get(study.patient_id)
            patient_name = f"{patient.first_name} {patient.last_name}" if patient else "N/A"

            status = []
            if study.is_pending:
                status.append("Pendiente")
            if study.is_critical:
                status.append("Crítico")
            if study.requires_followup:
                status.append("Seguimiento")

            ws.cell(row_idx, 1, study.study_date.strftime("%d/%m/%Y") if study.study_date else "")
            ws.cell(row_idx, 2, patient_name)
            ws.cell(row_idx, 3, study.study_type or "")
            ws.cell(row_idx, 4, study.study_name or "")
            ws.cell(row_idx, 5, study.institution or "")
            ws.cell(row_idx, 6, study.requesting_doctor or "")
            ws.cell(row_idx, 7, study.results or "")
            ws.cell(row_idx, 8, study.diagnosis or "")
            ws.cell(row_idx, 9, " | ".join(status))

            # Aplicar bordes
            for col in range(1, 10):
                cell = ws.cell(row_idx, col)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Resaltar críticos
                if study.is_critical:
                    cell.fill = critical_fill

        # Ajustar anchos
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 30
        ws.column_dimensions["I"].width = 15

        # Guardar
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
